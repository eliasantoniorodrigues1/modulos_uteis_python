# Documentação:
# https://openpyxl.readthedocs.io/en/stable/
# pip install openpyxl
import openpyxl
from random import uniform


pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_abas = pedidos.sheetnames
planilha1 = pedidos['Página1']

# print(planilha1['B4'].value)  # Pegar o valor
# for campo in planilha1['B']:
#     print(campo.value)  # Pegar a coluna inteira
#
# # Acessando um range:
# for linha in planilha1['A1:C2']:
#     for coluna in linha:
#         print(coluna.value)
#
#
# # For na planilha
# for linha in planilha1:
#     for coluna in linha:
#         print(coluna.value)

# Criar uma nova planilha:
# planilha1['B3'].value = 2200
# pedidos.save('nova_planilha.xlsx')

"""
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""

# Criar uma nova planilha
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']
